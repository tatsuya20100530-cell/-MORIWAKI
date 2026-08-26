#!/usr/bin/env python3
import csv
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
DATA = BASE / "data"
REPORTS = BASE / "reports"
REPORTS.mkdir(exist_ok=True)

JST = ZoneInfo("Asia/Tokyo")
NOW = datetime.now(JST)
TODAY = NOW.strftime("%Y-%m-%d")
YESTERDAY = (NOW.date() - timedelta(days=1)).isoformat()
CURRENT = DATA / f"{TODAY}.csv"
PREVIOUS = DATA / f"{YESTERDAY}.csv"
DAILY_XLSX = REPORTS / f"{TODAY}.xlsx"
LATEST_XLSX = REPORTS / "latest.xlsx"


def load_rows(path):
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def index_by_username(rows):
    return {r.get("username", ""): r for r in rows}


def delta(new, old, unit):
    try:
        d = int(new) - int(old)
        return f"{d:+d}{unit}"
    except Exception:
        return "算出不可"


def build_workbook():
    current = load_rows(CURRENT)
    previous = load_rows(PREVIOUS)
    prev = index_by_username(previous)

    wb = Workbook()
    ws = wb.active
    ws.title = "日次レポート"
    ws.freeze_panes = "A7"

    ws["A1"] = "MORIWAKI Instagram 日次チェック"
    ws["A1"].font = Font(size=18, bold=True)
    ws.merge_cells("A1:J1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"取得日時：{NOW.strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells("A2:J2")
    ws["A2"].alignment = Alignment(horizontal="center")

    success = sum(1 for r in current if r.get("result") == "OK")
    ws["A4"] = f"取得成功：{success}/20"
    ws["C4"] = f"比較基準：{YESTERDAY}" if PREVIOUS.exists() else f"比較基準：{YESTERDAY} 保存記録なし"

    headers = [
        "店舗", "スタッフ", "Instagram", "投稿数", "投稿前日比",
        "フォロワー数", "フォロワー前日比", "最新投稿日", "経過日数", "状態"
    ]
    header_row = 6
    for col, value in enumerate(headers, 1):
        cell = ws.cell(header_row, col, value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, r in enumerate(current, header_row + 1):
        old = prev.get(r.get("username", "")) if PREVIOUS.exists() else None
        if r.get("result") == "OK" and old and old.get("result") == "OK":
            post_diff = delta(r.get("posts"), old.get("posts"), "件")
            follower_diff = delta(r.get("followers"), old.get("followers"), "人")
        else:
            post_diff = "算出不可"
            follower_diff = "算出不可"

        values = [
            r.get("store", ""), r.get("staff", ""), r.get("username", ""),
            r.get("posts", "-"), post_diff, r.get("followers", "-"), follower_diff,
            r.get("latest", "-"), r.get("days", "-"), r.get("status", "")
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(i, col, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if post_diff not in ("+0件", "算出不可") or follower_diff not in ("+0人", "算出不可"):
            for col in (5, 7):
                ws.cell(i, col).font = Font(bold=True)

    widths = [13, 13, 26, 10, 14, 13, 18, 28, 11, 8]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[6].height = 28

    end = header_row + len(current)
    if end >= header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:J{end}"

    legend_row = end + 2
    ws.cell(legend_row, 1, "凡例：🟢 直近3日未満 / 🟡 3〜6日 / 🔴 7日以上 / ⚪ API取得不可")
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=10)
    ws.cell(legend_row, 1).alignment = Alignment(horizontal="left")

    rule_row = legend_row + 1
    ws.cell(rule_row, 1, "前日比：増加=+N、変動なし=+0、減少=-N。前日または当日の取得不可は算出不可。")
    ws.merge_cells(start_row=rule_row, start_column=1, end_row=rule_row, end_column=10)

    return wb


if not CURRENT.exists():
    raise SystemExit(f"Current CSV not found: {CURRENT}")

wb = build_workbook()
wb.save(DAILY_XLSX)
wb.save(LATEST_XLSX)
print(DAILY_XLSX)
print(LATEST_XLSX)
